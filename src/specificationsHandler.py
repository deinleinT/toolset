'''
Created on Apr 18, 2018

@author: deinlein thomas
'''

from openpyxl import load_workbook
import sys
import os
from lxml import html
import requests
import time
import threading
import zipfile
from win32com import client
from sys import exc_info


class SpecInfo():

    def __init__(self, spec, stype, title, status, primaryRespGrp, primaryRapporteur, initialPlannedRelease, publication, commonIMS, technology, important, versionOnlineOne, versionOnlineTwo, versionInXl, dateOnlineOne, dateOnlineTwo, dateInXl, urlOne, urlTwo, hyperlink, downloadOne, downloadTwo, notInOne, releaseOnlineOne, releaseOnlineTwo, versionInXlTwo, dateInXlTwo, versionOnlineThree, versionInXlThree, downloadThree, dateOnlineThree, dateInXlThree, urlThree, releaseOnlineThree, releaseInXLOne, releaseInXLTwo, releaseInXLThree):
        self.spec = spec
        self.title = title
        self.stype = stype
        self.versionOnlineOne = versionOnlineOne
        self.versionOnlineTwo = versionOnlineTwo
        self.versionOnlineThree = versionOnlineThree
        self.versionInXl = versionInXl
        self.versionInXlTwo = versionInXlTwo
        self.versionInXlThree = versionInXlThree
        self.dateOnlineOne = dateOnlineOne
        self.dateOnlineTwo = dateOnlineTwo
        self.dateOnlineThree = dateOnlineThree
        self.dateInXl = dateInXl
        self.dateInXlTwo = dateInXlTwo
        self.dateInXlThree = dateInXlThree
        self.urlOne = urlOne
        self.urlTwo = urlTwo
        self.urlThree = urlThree  
        self.status = status
        self.primaryRespGrp = primaryRespGrp
        self.primaryRapporteur = primaryRapporteur
        self.initialPlannedRelease = initialPlannedRelease
        self.publication = publication
        self.commonIMS = commonIMS
        self.technology = technology
        self.important = important
        self.hyperlink = hyperlink
        self.downloadOne = downloadOne
        self.downloadTwo = downloadTwo
        self.downloadThree = downloadThree
        self.notInOne = notInOne
        self.releaseOnlineOne = releaseOnlineOne
        self.releaseOnlineTwo = releaseOnlineTwo
        self.releaseOnlineThree = releaseOnlineThree
        self.releaseInXLOne = releaseOnlineOne
        self.releaseInXLTwo = releaseInXLTwo
        self.releaseInXLThree = releaseInXLThree


def currentUpdate(standardNumber, kind):
    
    print("Updating-Process of Specification-ID " + str(standardNumber) + " started...") 
    wb2 = load_workbook(sys.argv[2])
    sheetnames = wb2.sheetnames
    ws = wb2[sheetnames[0]]
    myIter = ws.iter_rows(row_offset=1)
    
    # Iterate over the rows in the Excel-sheet but skip the header.
    for row in myIter:
        number = row[0].value
        
        if str(number) == "None":
            continue
        if str(standardNumber) == str(number):
            if str(number) in failed:
                print("Do not update " + str(number) + " because extract failed!\n")
                continue
            if str(number) in str(noUpdate):
                print("Do not update " + str(number) + " because in noUpdate!\n")  
                continue  
            print("Updating the entry in the Excel-File of Specification-ID " + str(number))            
            if kind == "one":
                print("Update Version ONE")
                row[11].value = specs[str(number)].versionOnlineOne
                row[12].value = specs[str(number)].dateOnlineOne
                row[17].value = specs[str(number)].releaseOnlineOne
            elif kind == "two":
                print("Update Version TWO")
                row[13].value = specs[str(number)].versionOnlineTwo
                row[14].value = specs[str(number)].dateOnlineTwo
                row[18].value = specs[str(number)].releaseOnlineTwo
            elif kind == "three":
                print("Update Version THREE")
                row[15].value = specs[str(number)].versionOnlineThree
                row[16].value = specs[str(number)].dateOnlineThree
                row[19].value = specs[str(number)].releaseOnlineThree
        
    wb2.save(sys.argv[2])
    wb2.close()
    print("Updating-Process of Specification-ID " + str(standardNumber) + " finished!\n")


def getValues(folder):
    values = folder.split('_')[1]
    return values


def __convert_to_docx(doc, newName):
    
    print("Convert the file " + doc + " into DOCX.")
    flag = False
    path = os.getcwd()
    
    docTwo = path + doc.replace(".", "", 1).replace("/", "\\")
    newNameTwo = path + newName.replace(".", "", 1).replace("/", "\\")
    
    print("Converting " + str(docTwo) + " to " + str(newNameTwo))
    
    try:
        word = client.DispatchEx("Word.Application")  
        word.DisplayAlerts = False 
        word.Visible = False            
        worddoc = word.Documents.Open(docTwo)
        worddoc.SaveAs(newNameTwo, FileFormat=12)
        flag = True
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        flag = False
    finally:
        worddoc.Close(False)
        word.Quit()
        del word
        word = None
        if flag and (doc != newName):
            os.remove(doc)
            print("File " + str(doc) + " removed!")
        print("Was convert successful? --> " + str(flag))
        return flag


def __convert_to_pdf(doc, newName, delFlag):
    
    print("Convert the file " + doc + " into PDF.")
    flag = False
    path = os.getcwd()
    
    docTwo = path + doc.replace(".", "", 1).replace("/", "\\")
    newNameTwo = path + newName.replace(".", "", 1).replace("/", "\\")
    
    print("Converting " + str(docTwo) + " to " + str(newNameTwo))
    
    try:
        word = client.DispatchEx("Word.Application")  
        word.DisplayAlerts = False 
        word.Visible = False            
        worddoc = word.Documents.Open(docTwo)
        worddoc.SaveAs(newNameTwo, FileFormat=17)
        flag = True
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        flag = False
    finally:
        worddoc.Close(False)
        word.Quit()
        del word
        word = None
        if flag and delFlag:
            os.remove(doc)
            print("File " + str(doc) + " removed!")
        print("Was convert successful? --> " + str(flag))
        return flag
    

def extractZipFilesAndConvertToPDF(nameZip, destination, onlyWordFiles, convertDocx, convertDocxAndPDF):
    
    nameZipFile = destination + "/" + nameZip
    converted = False
    newName = nameZipFile.replace(".zip", ".pdf")
    newNameDocx = nameZipFile.replace(".zip", ".docx")
    standard = nameZip.split('_')[0]
    
    with zipfile.ZipFile(nameZipFile, "r") as zip_ref:
        docCounter = 0
        for name in zip_ref.namelist():
            if ".doc" in name or ".docx" in name:
                docCounter += 1
        
        if docCounter == 1:
            
            for name in zip_ref.namelist():
                if ".doc" in name or ".docx" in name:
                    docname = name
                    extracted = False
                    
                    try:
                        if ".doc"in name:
                            docname = nameZip.replace(".zip", ".doc")
                        if ".docx" in name:
                            docname = nameZip.replace(".zip", ".docx")                        
                        
                        zip_ref.extract(name, destination)
                        os.rename(destination + "/" + name, destination + "/" + docname)
                        print("Extracted the file " + docname)
                        extracted = True
                    except Exception as e:
                        logstring.append("ERROR in extractZipFilesMethod " + name + " ... NO Extract!!! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                    
                    print("Extraction successful? --> " + str(extracted))
                    if not extracted:
                        print("extracted Failed, hence add to noUpdate " + str(extracted))
                        failed.add(standard)
                        noUpdate.add(str(standard))
                    
                    if os.path.isfile(destination + "/" + docname):
                        if not onlyWordFiles and not convertDocx and not convertDocxAndPDF:
                            converted = __convert_to_pdf(destination + "/" + docname, newName, True)
                        if convertDocx:
                            converted = __convert_to_docx(destination + "/" + docname, newNameDocx)
                        if convertDocxAndPDF:
                            tempConverted = __convert_to_pdf(destination + "/" + docname, newName, False)
                            converted = __convert_to_docx(destination + "/" + docname, newNameDocx)
                            if not converted:
                                converted = tempConverted
                        if not converted:
                            if not extracted:
                                print("Extract and Convert Failed, hence add to noUpdate " + newName)
                                failed.add(standard)
                                noUpdate.add(str(standard))
                            else:
                                print("Extract: " + str(extracted) + " converted: " + str(converted) + " hence update ok, manual convert might be necessary!")

                else:
                    print("Not a Word-Document, skip " + name)
                    
        else:
            print("Did not extract " + nameZip + " (e.g., there is more than one file within the zip) but updated the entry in the Excel-Sheet!!!")
            return 
    if os.path.isfile(nameZipFile) and (converted or extracted):
        os.remove(nameZipFile) 
        print("Zip-File " + nameZipFile + " removed!")   
        return


def getAllFilesInDirectory(pathname, ending):
    dirList = os.listdir(pathname)
    dirList.sort()
    pdfs = []
    for g in dirList:
        if g.endswith(ending):
            pdfs.append(g)
    return pdfs


def extractAndConvert():
    pathname = "./Specifications"
    if not os.path.exists(pathname):
        os.makedirs(pathname)
        
    pdfs = []
    
    try:
        pdfs = getAllFilesInDirectory(pathname, ".zip")
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
    
    if len(pdfs) == 0:
        failed.clear()
        failed.add("-1")
        return 
    
    pdfcounter = 1
    for name in pdfs:
        standard = name.split('_')[0]
        print("***Progress of converting all Documents***: " + str(int(pdfcounter / len(pdfs) * 100)) + "%")
        print("Specification-ID is " + str(standard) + " and name of ZIP-File is " + name)
        if standard in noUpdate or standard in failed:
            print("Skip Extracting of " + str(name) + " because is in noUpdate or failed (e.g., Excel-File has already been updated).\n")
            continue
        try:
            print("Try to extract " + name)
            extractZipFilesAndConvertToPDF(name, pathname, onlyWordFiles, convertDocx, convertDocxAndPDF)
            print("Unzipping of " + name + " finished!\n")
        except Exception as e:
            logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        
        if str(specs[standard].downloadOne) == "1" and str(specs[standard].releaseOnlineOne) in name:
            currentUpdate(standard, "one")
        elif str(specs[standard].downloadTwo) == "1" and str(specs[standard].releaseOnlineTwo) in name:
            currentUpdate(standard, "two")
        elif str(specs[standard].downloadThree) == "1" and str(specs[standard].releaseOnlineThree) in name:
            currentUpdate(standard, "three")
        pdfcounter += 1           
    return 

    
def download(srcurl, dstfilepath):
 
    r = requests.get(srcurl)
    result = False

    if not os.path.exists("./Specifications/"):
        os.makedirs("./Specifications/")
        
    if not os.path.exists("./Specifications/" + dstfilepath):
        with open("./Specifications/" + dstfilepath, "wb") as code:
            code.write(r.content)
            result = True
    
    if os.path.exists("./Specifications/" + dstfilepath):
        result = True
            
    return result


def _getThreads():
    """ Returns the number of available threads on a posix/win based system """
    try:
        return (int)(os.environ['NUMBER_OF_PROCESSORS'])
    except:
        logstring.append("\n" + "Can not read number of processors... Terminate programm" + "\n")
        return -1
    
    
class fileTwo(threading.Thread):

    def __init__(self, name, check, start, end, noUpdate, specs):
        threading.Thread.__init__(self)
        self.__name = name
        self.__check = check
        self.__start = start
        self.__end = end
        self.__noUpdate = noUpdate
        self.__specs = specs
        
    def run(self):
        
        wb2 = load_workbook(self.__name)
        sheetnames = wb2.sheetnames
        ws = wb2[sheetnames[0]]
        
        myIter = None
        if self.__check:
            myIter = ws.iter_rows(row_offset=1, min_row=self.__start, max_row=self.__end)
        else:
            myIter = ws.iter_rows(min_row=self.__start, max_row=self.__end)
        
        print("\nProcessing Excel-File " + self.__name + " | " + str(threading.current_thread().getName()) + " | This Thread maintains the row-numbers " + str(self.__start) + " until " + str(self.__end) + "\n")
        counter = 1
        
        for row in myIter:
            specification = str(row[0].value)
            
            status = int(counter / (self.__end - self.__start) * 100)
            print(str(specification) + " \t\t| " + self.__name + " \t| " + str(threading.current_thread().getName()) + " \t| Status: " + str(status) + "%")
            counter += 1
            try:
                page = requests.get(self.__specs[specification].hyperlink)
                tree = html.fromstring(page.content)
            except Exception as e:
                logstring.append("\n" + "ERROR " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + " " + str(specification) + " " + self.__name + " " + str(threading.current_thread().getName()) + "\n")
                self.__noUpdate.add(str(specification))
                continue
            
            # version one
            for release in range(0, 1):
                    rowi = tree.xpath(
                        '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td/div/a'
                        .format(release))
                    
                    if len(rowi) > 0:
                        daterow = tree.xpath(
                            '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td'
                            .format(release))
                        self.__specs[specification].versionOnlineOne = "{}".format(rowi[1].text.strip())
                        
                        try:
                            rowReleaseOne = tree.xpath(
                                '//span[contains(@id,"SpecificationReleaseControl1_rpbReleases_i{}_HeaderTemplate_ctl00_lblReleaseName")]'
                                .format(release))
                            self.__specs[specification].releaseOnlineOne = rowReleaseOne[0].text.replace(' ', '')
                            
                        except Exception as e:
                            self.__specs[specification].releaseOnlineOne = ""
                            logstring.append("\n No Release-Version for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                         
                        try:
                            # for download
                            self.__specs[specification].urlOne = str(rowi[1].attrib["href"])
                            #
                        except Exception as e:
                            logstring.append("\n" + str(e) + " " + "ERROR During retrieving url for zipDownload " + str(specification) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                         
                        if daterow[2].text.strip() is not "":
                            date = daterow[2].text.split('-')
                            day = date[2].strip()
                            year = date[0].strip()
                            month = date[1].strip()
                            self.__specs[specification].dateOnlineOne = day + month + year
                        else:
                            self.__specs[specification].dateOnlineOne = "0"
                    else:
                        self.__specs[specification].versionOnlineOne = "0"
                        self.__specs[specification].urlOne = ""
                        self.__specs[specification].dateOnlineOne = "0"   

            # two
            for release in range(1, 2):
                    
                try:
                    try:
                        rowTwo = tree.xpath(
                        '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td/div/a'
                        .format(release))
                        dateRowTwo = tree.xpath(
                            '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td'
                            .format(release))
                        
                        try:
                            rowReleaseTwo = tree.xpath('//span[contains(@id,"SpecificationReleaseControl1_rpbReleases_i{}_HeaderTemplate_ctl00_lblReleaseName")]'.format(release))
                            self.__specs[specification].releaseOnlineTwo = rowReleaseTwo[0].text.replace(' ', '')
                        except Exception as e:
                            self.__specs[specification].releaseOnlineTwo = ""
                            logstring.append("\n No Release-Version for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                        
                        self.__specs[specification].versionOnlineTwo = "{}".format(rowTwo[1].text.strip())
                        if dateRowTwo[2].text.strip() is not "":
                            date = dateRowTwo[2].text.split('-')
                            day = date[2].strip()
                            year = date[0].strip()
                            month = date[1].strip()
                            self.__specs[specification].dateOnlineTwo = day + month + year
                        else:
                            self.__specs[specification].dateOnlineTwo = "0"
                        
                    except Exception as e:
                        logstring.append("\n No URL-Link for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                        
                    try:
                        self.__specs[specification].urlTwo = str(rowTwo[1].attrib["href"])
                        
                    except Exception as e:
                        logstring.append("\n No URL-Link for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                        
                except Exception as e:
                    logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
            # end two
            # three
            if checkThree:
                for release in range(2, 3):
                    try:
                        try:
                            rowThree = tree.xpath(
                            '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td/div/a'
                            .format(release))
                            dateRowThree = tree.xpath(
                                '//tr[@id="SpecificationReleaseControl1_rpbReleases_i{}_ctl00_specificationsVersionGrid_ctl00__0"]/td'
                                .format(release))
                            
                            try:
                                rowReleaseThree = tree.xpath('//span[contains(@id,"SpecificationReleaseControl1_rpbReleases_i{}_HeaderTemplate_ctl00_lblReleaseName")]'.format(release))
                                self.__specs[specification].releaseOnlineThree = rowReleaseThree[0].text.replace(' ', '')
                            except Exception as e:
                                self.__specs[specification].releaseOnlineThree = ""
                                logstring.append("\n No Release-Version for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                            
                            self.__specs[specification].versionOnlineThree = "{}".format(rowThree[1].text.strip())
                            if dateRowThree[2].text.strip() is not "":
                                date = dateRowThree[2].text.split('-')
                                day = date[2].strip()
                                year = date[0].strip()
                                month = date[1].strip()
                                self.__specs[specification].dateOnlineThree = day + month + year
                            else:
                                self.__specs[specification].dateOnlineThree = "0"
                            
                        except Exception as e:
                            logstring.append("\n No URL-Link for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                            
                        try:
                            self.__specs[specification].urlThree = str(rowThree[1].attrib["href"])
                            
                        except Exception as e:
                            logstring.append("\n No URL-Link for " + str(specification) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                            
                    except Exception as e:
                        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
                # end three
        print("\n\n\n")
        print("Processing of " + str(threading.current_thread().getName()) + " is finished!\n\n\n") 
        wb2.close()


class DownloadZIPs(threading.Thread):

    def __init__(self, specs, noUpdate, failed, kind):
        threading.Thread.__init__(self)
        self.__specs = specs
        self.__noUpdate = noUpdate
        self.__failed = failed
        self.__kind = kind
    
    def run(self):
        sortedKeys = sorted(self.__specs)      
    
        for key in sortedKeys:
            if self.__kind == "three" and self.__specs[key].downloadThree == "1" and checkThree:
                directory = self.__specs[key].spec + "_" + self.__specs[key].versionOnlineThree + "_" + self.__specs[key].dateOnlineThree + "_" + self.__specs[key].releaseOnlineThree + ".zip"
                try:
                    result = download(self.__specs[key].urlThree, directory)
                    if result:
                        print("File " + directory + " created!\n")  
                    else:
                        self.__failed.add(str(key)) 
                        print("Download file " + directory + " FAILED!\n")                 
                except Exception as e:
                    self.__failed.add(str(key)) 
                    logstring.append("\n" + str(e) + " " + "Download file " + directory + " FAILED!\n" + "ERROR in downloadZIpsRUN" + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
           
            elif self.__kind == "two" and self.__specs[key].downloadTwo == "1":
                directory = self.__specs[key].spec + "_" + self.__specs[key].versionOnlineTwo + "_" + self.__specs[key].dateOnlineTwo + "_" + self.__specs[key].releaseOnlineTwo + ".zip"
                try:
                    result = download(self.__specs[key].urlTwo, directory)
                    if result:
                        print("File " + directory + " created!\n")  
                    else:
                        self.__failed.add(str(key)) 
                        print("Download file " + directory + " FAILED!\n")                 
                except Exception as e:
                    self.__failed.add(str(key)) 
                    logstring.append("\n" + str(e) + " " + "Download file " + directory + " FAILED!\n" + "ERROR in downloadZIpsRUN" + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")

            elif self.__kind == "one" and self.__specs[key].downloadOne == "1":
                directory = self.__specs[key].spec + "_" + self.__specs[key].versionOnlineOne + "_" + self.__specs[key].dateOnlineOne + "_" + self.__specs[key].releaseOnlineOne + ".zip"
                
                try:
                    result = download(self.__specs[key].urlOne, directory)
                    if result:
                        print("File " + directory + " created!\n")  
                    else:
                        self.__failed.add(str(key)) 
                        print("Download file " + directory + " FAILED!\n")                 
                except Exception as e:
                    self.__failed.add(str(key)) 
                    logstring.append("\n" + str(e) + " " + "Download file " + directory + " FAILED!\n" + "ERROR in downloadZIpsRUN" + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")


#############################################################################################
#################Programm Start##############################################################
logstring = []

##################################################################################################################################################################
checkThree = False
if "-3" in sys.argv:
    checkThree = True
####################################################################################################################################################################
onlyWordFiles = False
if "-w" in sys.argv:
    onlyWordFiles = True
#####################################################################################################################################################################
convertDocx = False
if "-wx" in sys.argv:
    convertDocx = True
    
###
convertDocxAndPDF = False
if "-wxp" in sys.argv:
    convertDocxAndPDF = True
    convertDocx = False
###

if onlyWordFiles:
    convertDocx = False
    convertDocxAndPDF = False
#####################################################################################################################################################################

specs = dict()
noUpdate = set()
datumDifference = "checkSpecificationsBetween" + sys.argv[1] + "And" + sys.argv[2] + time.strftime("%d%m%Y%H%M") + "_"
logFileName = "LogFile" + sys.argv[1] + "And" + sys.argv[2] + time.strftime("%d%m%Y%H%M") + "_"

#############################################################################################
threadLock = threading.Lock()
threads = []
numberThreads = _getThreads() * 2
print("\n*** Using " + str(numberThreads) + " Threads for comparing the versions of all specifications in the Excel-Sheets... ***\n")
##################################################
#
wbOne = load_workbook(sys.argv[1])
sheetnames = wbOne.sheetnames
wsOne = wbOne[sheetnames[0]]
standardsInOne = []
iterOne = wsOne.iter_rows(row_offset=1, min_row=0, max_row=wsOne.max_row)
for row in iterOne:
    try:
        standardsInOne.append(str(row[0].value))
        specs[str(row[0].value)] = SpecInfo(spec=str(row[0].value),
                                            stype="" if str(row[1].value) == "None" else str(row[1].value),
                                            title="" if str(row[2].value) == "None" else str(row[2].value),
                                            status="" if str(row[3].value) == "None" else str(row[3].value),
                                            primaryRespGrp="" if str(row[4].value) == "None" else str(row[4].value),
                                            primaryRapporteur="" if str(row[5].value) == "None" else str(row[5].value),
                                            initialPlannedRelease="" if str(row[6].value) == "None" else str(row[6].value),
                                            publication="" if str(row[7].value) == "None" else str(row[7].value),
                                            commonIMS="" if str(row[8].value) == "None" else str(row[8].value),
                                            technology="" if str(row[9].value) == "None" else str(row[9].value),
                                            hyperlink="" if str(row[0].value) == "None" else str(row[0].hyperlink.target),
                                            important="",
                                            versionOnlineOne="0",
                                            versionOnlineTwo="0",
                                            versionOnlineThree="0",
                                            versionInXl="0",
                                            dateOnlineOne="0",
                                            dateOnlineTwo="0",
                                            dateOnlineThree="0",
                                            dateInXl="0",
                                            urlOne="",
                                            urlTwo="",
                                            urlThree="",
                                            downloadOne="0",
                                            downloadTwo="0",
                                            downloadThree="0",
                                            notInOne="0",
                                            releaseOnlineOne="",
                                            releaseOnlineTwo="",
                                            releaseOnlineThree="",
                                            versionInXlTwo="0",
                                            versionInXlThree="0",
                                            dateInXlTwo="0",
                                            dateInXlThree="0",
                                            releaseInXLOne="0",
                                            releaseInXLTwo="0",
                                            releaseInXLThree="0")
    except Exception as e:
        logstring.append("\n No URL-Link for " + str(row[0].value) + " available! " + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")

specs.pop('None', None)
wbTwo = load_workbook(sys.argv[2])
sheetnames = wbTwo.sheetnames
wsTwo = wbTwo[sheetnames[0]]
standardsInTwo = []
iterTwo = wsTwo.iter_rows(row_offset=1, min_row=0, max_row=wsTwo.max_row)
for row in iterTwo:
    try:
        number = str(row[0].value)
        standardsInTwo.append(number)
        if number not in standardsInOne:
            specs[str(row[0].value)] = SpecInfo(spec=str(row[0].value),
                                            stype="" if str(row[1].value) == "None" else str(row[1].value),
                                            title="" if str(row[2].value) == "None" else str(row[2].value),
                                            status="" if str(row[3].value) == "None" else str(row[3].value),
                                            primaryRespGrp="" if str(row[4].value) == "None" else str(row[4].value),
                                            primaryRapporteur="" if str(row[5].value) == "None" else str(row[5].value),
                                            initialPlannedRelease="" if str(row[6].value) == "None" else str(row[6].value),
                                            publication="" if str(row[7].value) == "None" else str(row[7].value),
                                            commonIMS="" if str(row[8].value) == "None" else str(row[8].value),
                                            technology="" if str(row[9].value) == "None" else str(row[9].value),
                                            hyperlink="" if str(row[0].value) == "None" else str(row[0].hyperlink.target),
                                            important="" if str(row[10].value) == "None" else str(row[10].value),
                                            versionOnlineOne="0",
                                            versionOnlineTwo="0",
                                            versionOnlineThree="0",
                                            versionInXl="0" if str(row[11].value) == "None" else str(row[11].value),
                                            dateOnlineOne="0",
                                            dateOnlineTwo="0",
                                            dateOnlineThree="0",
                                            dateInXl="0" if str(row[12].value) == "None" else str(row[12].value),
                                            urlOne="",
                                            urlTwo="",
                                            urlThree="",
                                            downloadOne="0",
                                            downloadTwo="0",
                                            downloadThree="0",
                                            notInOne="1",
                                            releaseOnlineOne="",
                                            releaseOnlineTwo="",
                                            releaseOnlineThree="",
                                            versionInXlTwo="0" if str(row[13].value) == "None" else str(row[13].value),
                                            versionInXlThree="0" if str(row[15].value) == "None" else str(row[15].value),
                                            dateInXlTwo="0" if str(row[14].value) == "None" else str(row[14].value),
                                            dateInXlThree="0" if str(row[16].value) == "None" else str(row[16].value),
                                            releaseInXLOne="0" if str(row[17].value) == "None" else str(row[17].value),
                                            releaseInXLTwo="0" if str(row[18].value) == "None" else str(row[18].value),
                                            releaseInXLThree="0" if str(row[19].value) == "None" else str(row[19].value))
        else:
            specs[number].important = "" if str(row[10].value) == "None" else str(row[10].value)
            specs[number].versionInXl = "0" if str(row[11].value) == "None" else str(row[11].value)
            specs[number].dateInXl = "0" if str(row[12].value) == "None" else str(row[12].value)
            specs[number].versionInXlTwo = "0" if str(row[13].value) == "None" else str(row[13].value)
            specs[number].dateInXlTwo = "0" if str(row[14].value) == "None" else str(row[14].value)
            specs[number].versionInXlThree = "0" if str(row[15].value) == "None" else str(row[15].value)
            specs[number].dateInXlThree = "0" if str(row[16].value) == "None" else str(row[16].value)
            specs[number].releaseInXLOne = "0" if str(row[17].value) == "None" else str(row[17].value)
            specs[number].releaseInXLTwo = "0" if str(row[18].value) == "None" else str(row[18].value)
            specs[number].releaseInXLThree = "0" if str(row[19].value) == "None" else str(row[19].value)
                        
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")

rowCounter = wsTwo.max_row
for specification in standardsInOne:
    if specification not in standardsInTwo:
        try:
            wsTwo.cell(row=rowCounter, column=1).value = specification
            wsTwo.cell(row=rowCounter, column=1).hyperlink = specs[specification].hyperlink
            wsTwo.cell(row=rowCounter, column=2).value = specs[specification].stype
            wsTwo.cell(row=rowCounter, column=3).value = specs[specification].title
            wsTwo.cell(row=rowCounter, column=4).value = specs[specification].status
            wsTwo.cell(row=rowCounter, column=5).value = specs[specification].primaryRespGrp
            wsTwo.cell(row=rowCounter, column=6).value = specs[specification].primaryRapporteur
            wsTwo.cell(row=rowCounter, column=7).value = specs[specification].initialPlannedRelease
            wsTwo.cell(row=rowCounter, column=8).value = specs[specification].publication
            wsTwo.cell(row=rowCounter, column=9).value = specs[specification].commonIMS
            wsTwo.cell(row=rowCounter, column=10).value = specs[specification].technology
            wsTwo.cell(row=rowCounter, column=11).value = specs[specification].important
            wsTwo.cell(row=rowCounter, column=12).value = specs[specification].versionInXl
            wsTwo.cell(row=rowCounter, column=13).value = specs[specification].dateInXl
            wsTwo.cell(row=rowCounter, column=14).value = specs[specification].versionInXlTwo
            wsTwo.cell(row=rowCounter, column=15).value = specs[specification].dateInXlTwo
            wsTwo.cell(row=rowCounter, column=16).value = specs[specification].versionInXlThree
            wsTwo.cell(row=rowCounter, column=17).value = specs[specification].dateInXlThree
            wsTwo.cell(row=rowCounter, column=18).value = specs[specification].releaseInXLOne
            wsTwo.cell(row=rowCounter, column=19).value = specs[specification].releaseInXLTwo
            wsTwo.cell(row=rowCounter, column=20).value = specs[specification].releaseInXLThree
            rowCounter += 1
        except Exception as e:
            logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + " " + str(specification) + "\n")
            
wbTwo.save(sys.argv[2])
wbOne.close()
wbTwo.close()
standardsInOne.clear()
standardsInTwo.clear()
##################################################

wb2 = load_workbook(sys.argv[1])
sheetnames = wb2.sheetnames
ws = wb2[sheetnames[0]]
maxFirst = ws.max_row
wb2.close()

wb3 = load_workbook(sys.argv[2])
sheetnames = wb3.sheetnames
ws = wb3[sheetnames[0]]
maxSecond = ws.max_row

myIter = ws.iter_rows(row_offset=1, min_row=0, max_row=ws.max_row)
nextNr = "0"
impCol = 11
versionCol = 12
dateCol = 13
versionCol2 = 14
dateCol2 = 15
versionCol3 = 16
dateCol3 = 17
relCol = 18
relCol2 = 19
relCol3 = 20
changeSecondExcelSheet = False
counter = 0

ws.cell(row=1, column=impCol).value = "Important"
ws.cell(row=1, column=versionCol).value = "Version"
ws.cell(row=1, column=dateCol).value = "Date"
ws.cell(row=1, column=versionCol2).value = "VersionTwo"
ws.cell(row=1, column=dateCol2).value = "DateTwo"
ws.cell(row=1, column=versionCol3).value = "VersionThree"
ws.cell(row=1, column=dateCol3).value = "DateThree"
ws.cell(row=1, column=relCol).value = "ReleaseOne"
ws.cell(row=1, column=relCol2).value = "ReleaseTwo"
ws.cell(row=1, column=relCol3).value = "ReleaseThree"
wb3.save(sys.argv[2])
myIter = ws.iter_rows(row_offset=1, min_row=0, max_row=ws.max_row)

for row in myIter:
    currentRow = row[0].row
    specification = str(row[0].value)
    if str(row[0].value) == "None":
        continue
    
    try:
        if str(row[1].value) == "None" or str(row[1].value) == "":
            row[1].value = specs[specification].stype
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=2).value = ""
        
    try:
        if str(row[2].value) == "None" or str(row[2].value) == "":
            row[2].value = specs[specification].title
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=3).value = ""
        
    try:
        if str(row[3].value) == "None" or str(row[3].value) == "":
            row[3].value = specs[specification].status
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=4).value = ""
        
    try:
        if str(row[4].value) == "None" or str(row[4].value) == "":
            row[4].value = specs[specification].primaryRespGrp
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=5).value = ""
        
    try:
        if str(row[5].value) == "None" or str(row[5].value) == "":
            row[5].value = specs[specification].primaryRapporteur
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=6).value = ""
        
    try:
        if str(row[6].value) == "None" or str(row[6].value) == "":
            row[6].value = specs[specification].initialPlannedRelease
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=7).value = ""
    
    try:
        if str(row[7].value) == "None" or str(row[7].value) == "":
            row[7].value = specs[specification].publication
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=8).value = ""
    
    try:
        if str(row[8].value) == "None" or str(row[8].value) == "":
            row[8].value = specs[specification].commonIMS
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=9).value = ""
    
    try:
        if str(row[9].value) == "None" or str(row[9].value) == "":
            row[9].value = specs[specification].technology
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=10).value = ""

    try:
        if str(row[10].value) == "None" or str(row[10].value) == "":
            row[10].value = ""
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=11).value = ""
            
    try:
        if str(row[11].value) == "None" or str(row[11].value) == "":
            row[11].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=12).value = "0"
        
    try:
        if str(row[12].value) == "None" or str(row[12].value) == "":
            row[12].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=13).value = "0"
    
    try:
        if str(row[13].value) == "None" or str(row[13].value) == "":
            row[13].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=14).value = "0"
    
    try:
        if str(row[14].value) == "None" or str(row[14].value) == "":
            row[14].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=15).value = "0"
        
    try:
        if str(row[15].value) == "None" or str(row[15].value) == "":
            row[15].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=16).value = "0"
    
    try:
        if str(row[16].value) == "None" or str(row[16].value) == "":
            row[16].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=17).value = "0"
    
    try:
        if str(row[17].value) == "None" or str(row[17].value) == "":
            row[17].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=18).value = "0"
    
    try:
        if str(row[18].value) == "None" or str(row[18].value) == "":
            row[18].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=19).value = "0"
    
    try:
        if str(row[19].value) == "None" or str(row[19].value) == "":
            row[19].value = "0"
    except Exception as e:
        logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
        ws.cell(row=currentRow, column=20).value = "0"

wb3.save(sys.argv[2])
wb3.close()

#
# calc the borders
print("*** Starting all Threads *** \n\n")
startOne = 0
startTwo = 0
intervalTwo = round(maxSecond / numberThreads)
if intervalTwo == 0:
    intervalTwo = 1;

thread2 = fileTwo(sys.argv[2], True, startTwo, startTwo + intervalTwo, noUpdate, specs)
startTwo += intervalTwo

threads.append(thread2)

while startTwo < maxSecond:
    endTwo = startTwo + intervalTwo
    if endTwo > maxSecond:
        endTwo = maxSecond
    threads.append(fileTwo(sys.argv[2], False, startTwo, endTwo, noUpdate, specs))  
    startTwo = endTwo 

for t in threads:
    t.start()

for t in threads:
    t.join()
print("\n*******Threads finished!*********\n")

#############################################################################################
###############################compare version online and in ExcelSheet######################
standardsToCheck = []
outputString = []
failed = set()

for number in sorted(specs):
    number = str(number)
    versionInXl = specs[number].versionInXl
    versionInXlTwo = specs[number].versionInXlTwo
    versionInXlThree = specs[number].versionInXlThree
    versionOnlineOne = specs[number].versionOnlineOne
    versionOnlineTwo = specs[number].versionOnlineTwo
    versionOnlineThree = specs[number].versionOnlineThree
    dateInXl = specs[number].dateInXl
    dateInXlTwo = specs[number].dateInXlTwo
    dateInXlThree = specs[number].dateInXlThree
    dateOnlineOne = specs[number].dateOnlineOne
    dateOnlineTwo = specs[number].dateOnlineTwo
    dateOnlineThree = specs[number].dateOnlineThree
    title = specs[number].title
    important = specs[number].important
    notInOne = specs[number].notInOne
    releaseOnlineOne = specs[number].releaseOnlineOne
    releaseOnlineTwo = specs[number].releaseOnlineTwo
    releaseOnlineThree = specs[number].releaseOnlineThree
    releaseInXLOne = specs[number].releaseInXLOne
    releaseInXLTwo = specs[number].releaseInXLTwo
    releaseInXLThree = specs[number].releaseInXLThree
    urlOne = specs[number].urlOne
    urlTwo = specs[number].urlTwo
    urlThree = specs[number].urlThree

    if notInOne == "1":
        outputString.append(str(number) + " not in " + sys.argv[1] + "\n")
        standardsToCheck.append(str(number))
    
    if versionInXl == versionOnlineOne and dateInXl == dateOnlineOne:
        pass
    else:
        if versionOnlineOne == "0" or dateOnlineOne == "0" or dateOnlineOne == "None" or dateOnlineOne == "" or versionOnlineOne == "None" or versionOnlineOne == "":
            outputString.append("\n*** Check MANUALLY Version One!!! ***")
            outputString.append("" + releaseOnlineOne)
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Version than Online! " + str(versionInXl) + " != " + str(versionOnlineOne))
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Date than Online! " + str(dateInXl) + " != " + str(dateOnlineOne))
            outputString.append("***\n")
            standardsToCheck.append(number)
        else:
            outputString.append("" + releaseOnlineOne)
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Version than Online! " + str(versionInXl) + " != " + str(versionOnlineOne))
            if number not in standardsToCheck:
                standardsToCheck.append(number)
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Date than Online! " + str(dateInXl) + " != " + str(dateOnlineOne) + "\n")
            try:
                if specs[number].urlOne != "None" and specs[number].urlOne != "":
                    specs[number].downloadOne = "1"
            except Exception as e:
                logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "Error adding key " + str(number) + " to downloadZips" + "\n")
                noUpdate.add(str(number))
    
    if versionInXlTwo == versionOnlineTwo and dateInXlTwo == dateOnlineTwo:
        pass
    else:
        if versionOnlineTwo == "0" or dateOnlineTwo == "0" or versionOnlineTwo == "None" or dateOnlineTwo == "" or versionOnlineTwo == "None" or dateOnlineTwo == "":
            outputString.append("\n*** Check MANUALLY Version Two!!! ***")
            outputString.append("" + releaseOnlineTwo)
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Version than Online! " + str(versionInXlTwo) + " != " + str(versionOnlineTwo))
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Date than Online! " + str(dateInXlTwo) + " != " + str(dateOnlineTwo))
            outputString.append("***\n")
            standardsToCheck.append(number)
        else:
            outputString.append("" + releaseOnlineTwo)
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Version than Online! " + str(versionInXlTwo) + " != " + str(versionOnlineTwo))
            if number not in standardsToCheck:
                standardsToCheck.append(number)
            outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Date than Online! " + str(dateInXlTwo) + " != " + str(dateOnlineTwo) + "\n")
            try:
                if specs[number].urlTwo != "None" and specs[number].urlTwo != "":
                    specs[number].downloadTwo = "1"
            except Exception as e:
                logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "Error adding key " + str(number) + " to downloadZips" + "\n")
                noUpdate.add(str(number))
                
    if checkThree:
        if versionInXlThree == versionOnlineThree and dateInXlThree == dateOnlineThree:
            pass
        else:
            if versionOnlineThree == "0" or dateOnlineThree == "0" or versionOnlineThree == "None" or dateOnlineThree == "" or versionOnlineThree == "None" or dateOnlineThree == "":
                outputString.append("\n*** Check MANUALLY Version Three!!! ***")
                outputString.append("" + releaseOnlineThree)
                outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Version than Online! " + str(versionInXlThree) + " != " + str(versionOnlineThree))
                outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Date than Online! " + str(dateInXlThree) + " != " + str(dateOnlineThree))
                outputString.append("***\n")
                standardsToCheck.append(number)
            else:
                outputString.append("" + releaseOnlineThree)
                outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Version than Online! " + str(versionInXlThree) + " != " + str(versionOnlineThree))
                if number not in standardsToCheck:
                    standardsToCheck.append(number)
                outputString.append(str(number) + " in " + str(sys.argv[2]) + " has different Date than Online! " + str(dateInXlThree) + " != " + str(dateOnlineThree) + "\n")
                try:
                    if specs[number].urlThree != "None" and specs[number].urlThree != "":
                        specs[number].downloadThree = "1"
                except Exception as e:
                    logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "Error adding key " + str(number) + " to downloadZips" + "\n")
                    noUpdate.add(str(number))
 
datumDifference += time.strftime("%d%m%Y%H%M") + ".txt"
logFileName += time.strftime("%d%m%Y%H%M") + ".txt"
out = open(datumDifference, "w")
out.write("\nIMPORTANT Specifications with changes: \n")      
for r in standardsToCheck:
    if str(r) in specs.keys():
        if(specs[r].important == "1"):
            out.write(str(r) + " " + specs[str(r)].title + "\n")
out.write("\n\nSpecification with changes or differences: \n")
for r in standardsToCheck:
    if str(r) in specs.keys():
        out.write(str(r) + " " + specs[str(r)].title + "\n")
  
out.write("\n\n\n\nThe differences in detail:\n")
for r in outputString:
    out.write(str(r))
    out.write("\n")
out.close()
print("\n\n****************************File " + datumDifference + " saved!************************************************\n\n")

print("\nStarting Saving the zips\n")
t1 = DownloadZIPs(specs, noUpdate, failed, "one")
t2 = DownloadZIPs(specs, noUpdate, failed , "two")

if checkThree:
    t3 = DownloadZIPs(specs, noUpdate, failed, "three")
    thAr = [t1, t2, t3]
else:
    thAr = [t1, t2]
for t in thAr:
    t.start()

for t in thAr:
    t.join()

print("\n*****************Finished Saving ZIPS*********************")
###############################################################################################################################
print("\n\n\n*****************Extracting Zips and Convert*********************")

print("\nStart...\n")
try:
    extractAndConvert()
except Exception as e:
    logstring.append("\n" + str(e) + " " + str(exc_info()) + " LineNumber: " + str(sys._getframe().f_lineno) + "\n")
    
print("\n\n*********************Finished extracting and converting********************************\n\n")
################################################################################################################################

print("\n\n\n*****************Writing LogFile*********************")

out = open(logFileName, "w")
out.write("\n")
out.write("\n\nLogFile:\n")
for s in logstring:
    out.write(s)
out.close()

print("\n\n\n*****************FINISHED*********************")

