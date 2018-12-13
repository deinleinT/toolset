'''
Created on Apr 18, 2018

@author: deinlein thomas
'''

from openpyxl import load_workbook
from bibtexparser.bwriter import BibTexWriter
from bibtexparser.bibdatabase import BibDatabase
import sys


def toBibDatabase(number, version, date, release, url, title, types, dbArray):
    if number == "0" or version == "0" or date == "0" or url == "None":    
        return
    entry = {
            'ID': str(number + "V" + version + "D" + date),
            'ENTRYTYPE': "techreport",
            'title': title,
            'type': types,
            'author': "{3rd Generation Partnership Project (3GPP)}",
            'number': number,
            'note': __convertReleaseString(release) + ", Version: " + version + ", Published: " + __convertDateString(date),
            'url': url
        }
    print("Bib-Entry created for Specification " + str(number + "V" + version + "D" + date))
    dbArray.append(entry)


def __convertReleaseString(releasestring):
    rel = releasestring[0:7]
    num = releasestring[7:]
    return rel + " " + num


def __convertDateString(datestring):
    year = datestring[4:]
    month = datestring[2:4] 
    day = datestring[0:2]
    return year + "-" + month + "-" + day
    

db = BibDatabase()
db.entries = []
wb2 = load_workbook(sys.argv[1])
sheetnames = wb2.sheetnames
ws = wb2[sheetnames[0]]
filename = sys.argv[2]

# Iterate over the rows in the Excel-sheet but skip the header.
for row in ws.iter_rows(row_offset=1):
    number = str(row[0].value)
    if number == "None":
        continue
    types = str(row[1].value)
    title = str(row[2].value)
    important = str(row[10].value)
    versionOne = str(row[11].value)
    dateOne = str(row[12].value)
    versionTwo = str(row[13].value)
    dateTwo = str(row[14].value)
    versionThree = str(row[15].value)
    dateThree = str(row[16].value)
    releaseOne = str(row[17].value)
    releaseTwo = str(row[18].value)
    releaseThree = str(row[19].value)
    
    try:
        url = str(row[0].hyperlink.target)
    except:
        url = "None"
    
    try:
        if sys.argv[3] == "-i":
            if important != "1":
                continue
    except:
        pass

    toBibDatabase(number, versionOne, dateOne, releaseOne, url, title, types, db.entries)
    toBibDatabase(number, versionTwo, dateTwo, releaseTwo, url, title, types, db.entries)
    toBibDatabase(number, versionThree, dateThree, releaseThree, url, title, types, db.entries)
writer = BibTexWriter()
with open(filename, 'w') as bibfile:
    bibfile.write(writer.write(db))
