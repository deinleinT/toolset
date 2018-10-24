from openpyxl import load_workbook
from bibtexparser.bwriter import BibTexWriter
from bibtexparser.bibdatabase import BibDatabase
import sys


def __convertDateString(datestring):
    year = datestring[4:]
    month = datestring[2:4] 
    day = datestring[0:2]
    return year + "-" + month + "-" + day
    

db = BibDatabase()
db.entries = []
wb2 = load_workbook(sys.argv[1])
sheetnames = wb2.sheetnames
ws = wb2[sheetnames[0]]
filename = sys.argv[2]

# Iterate over the rows in the Excel-sheet but skip the header.

for row in ws.iter_rows(row_offset=1):
    number = str(row[0].value)
    if number == "None":
        continue
    types = str(row[1].value)
    title = str(row[2].value)
    important = str(row[10].value)
    versionOne = str(row[11].value)
    dateOne = str(row[12].value)
    versionTwo = str(row[13].value)
    dateTwo = str(row[14].value)
    try:
        url = str(row[0].hyperlink.target)
    except:
        url = "None"
    
    try:
        if sys.argv[3] == "-i":
            if important != "1":
                continue
    except:
        pass

# version one
    if number == "0" or versionOne == "0" or dateOne == "0" or url == "None":
        pass
    else:
        entry = {
            'ID': str(number + "V" + versionOne + "D" + dateOne),
            'ENTRYTYPE': "techreport",
            'title': title,
            'type': types,
            'author': "{3rd Generation Partnership Project (3GPP)}",
            'number': number,
            'note': versionOne + ", " + __convertDateString(dateOne),
            'url': url
        }
        print("Bib-Entry created for Specification " + str(number + "V" + versionOne + "D" + dateOne))
        db.entries.append(entry)
    
# version tow
    if number == "0" or versionTwo == "0" or dateTwo == "0" or url == "None":
        pass
    else:
        entry = {
            'ID': str(number + "V" + versionTwo + "D" + dateTwo),
            'ENTRYTYPE': "techreport",
            'title': title,
            'type': types,
            'author': "{3rd Generation Partnership Project (3GPP)}",
            'number': number,
            'note': versionTwo + ", " + __convertDateString(dateTwo),
            'url': url
        }
        print("Bib-Entry created for Specification " + str(number + "V" + versionTwo + "D" + dateTwo))
        db.entries.append(entry)

writer = BibTexWriter()
with open(filename, 'w') as bibfile:
    bibfile.write(writer.write(db))
